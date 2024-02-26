"""This script is intended to be run as a one-off to generate the v1 truth set from manually curated evidence.

The manually curated evidence is stored in a spreadsheed in google sheets, and will need to be manually downloaded
to a temporary location before running this script.

Note: we're not currently handling authentication to google sheets, so the spreadsheet will need to be manually
downloaded and uploaded to Azure blob storage before processing.

Note: because this is a the spreadsheet contains multiple sheets, it will be necessary to download it as an xlsx file,
not a csv or tsv file.

TODO (2/21/2024) -
- We're not currently validating whether any of the variants specified in the truth set are syntatically or biologically
    correct.
- Along similar lines, the validators don't currently handle p. or c. notation.
- Probably not necessary, but we don't validate the HPO or OMIM terms to make sure they exist.
- Train and test are just taken as an fractional split of genes per evidence category. They will be re-selected any
    time this script is re-run. Fixed split is probably more appropriate.

"""

# TODO:
# Functional_study is obsolete, traded for three subsequent CONTENT_SHEET_COLUMN_HEADINGS

# %% Imports.


import os
import re
from functools import cache
from typing import Any, List, Tuple

import openpyxl
import pandas as pd

from lib.evagg.ref import NcbiLookupClient
from lib.evagg.svc import CosmosCachingWebClient, get_dotenv_settings

# %% Constants.

TRUTH_XLSX = "/home/azureuser/repos/ev-agg-exp/.tmp/Manual Ground Truth.xlsx"
OUTPUT_ROOT = "/home/azureuser/repos/ev-agg-exp/data/v1"

TRAIN_FRAC = 0.7

PAPERS_SHEET_NAME = "find_right_papers"

CONTENT_SHEET_COLUMN_HEADINGS = {
    "paper_build": "paper_build",
    "paper_variant": "paper_variant",
    "Individual_ID": "individual_id",
    "Text Description": "pheno_text_description",
    "Phenotype": "phenotype",
    "Transcript": "transcript",
    "HGVS.C": "hgvs_c",
    "HGVS.P": "hgvs_p",
    "Variant_Type": "variant_type",
    "Zygosity": "zygosity",
    "Variant_Inheritance": "variant_inheritance",
    "Paper_ID (PMID)": "pmid",
    "Author (first auth. last name)": "author",
    "Year": "year",
    "Journal (full name)": "journal",
    "Study_Type": "study_type",
    "Functional_Study": "functional_study",
    "In_Supplement": "in_supplement",
    "Notes": "notes",
    "Questions": "questions",
}

FORCED_STRING_FIELDS = ["pmid", "individual_id", "year"]

# %% Content validators.


def _paper_build_validator(x: Any) -> Tuple[bool, str]:
    result = x in ["GRCh37", "hg19", "GRCh38", "hg38", "T2T", "Unknown", None]
    return result, "Value is not one of GRCh37/hg19, GRCh38/hg38, or None" if not result else ""


def _not_none_validator(x: Any) -> Tuple[bool, str]:
    result = x is not None
    return result, "Value is None" if not result else ""


def _phenotype_validator(x: Any) -> Tuple[bool, str]:
    # Phenotype should be a comma-delimited list of, each one is terminated by
    # the string "(HPO:\d+)"" or "(OMIM:\d+)".
    # Alternatively, it can be Unknown.
    result = x is not None and (
        x == "Unknown" or all(re.search(r"\(HP:\d+\)$|\(OMIM:\d+\)$", phenotype.strip()) for phenotype in x.split(","))
    )
    return result, "Value is not a semicolon-delimited list of HPO or OMIM terms" if not result else ""


def _transcript_validator(x: Any) -> Tuple[bool, str]:
    result = x is None or x.startswith("NM_") or x.startswith("ENST")  # Transcript is optional.
    return result, "Value must be None or start with NM_ or ENST" if not result else ""


def _hgvs_c_validator(x: Any) -> Tuple[bool, str]:
    # TODO, not particularly sophisticated.
    result = x is None or x.startswith("c.")
    return result, "Value must be None or start with 'c.'" if not result else ""


def _hgvs_p_validator(x: Any) -> Tuple[bool, str]:
    # TODO, not particularly sophisticated.
    result = x is None or x.startswith("p.")  # hgvs.p is optional
    return result, "Value must be None or start with 'p.'" if not result else ""


def _is_digit_validator(x: Any) -> Tuple[bool, str]:
    result = x is not None and str(x).replace(".", "").isdigit()
    return result, "Value is not a digit" if not result else ""


_VARIANT_TYPES = [
    "missense",
    "frameshift",
    "stop gained",
    "splice donor",
    "splice acceptor",
    "splice region",
    "start lost",
    "inframe deletion",
    "frameshift deletion",
    "inframe insertion",
    "frameshift insertion",
    "structural",
    "synonymous",
    "intron",
    "5' UTR",
    "3' UTR",
    "non-coding",
    "unknown",
]


def _variant_type_validator(x: Any) -> Tuple[bool, str]:
    result = x in _VARIANT_TYPES
    return result, f"Value is not one of {_VARIANT_TYPES}" if not result else ""


_ZYGOSITIES = [
    "heterozygous",
    "homozygous",
    "compound heterozygous",
    "none",
    "heterozygous AD",
    "heterozygous AR",
    "heterozygous XLD",
    "heterozygous XLR",
]


def _zygosity_validator(x: Any) -> Tuple[bool, str]:
    result = x in _ZYGOSITIES
    return result, f"Value is not one of {_ZYGOSITIES}" if not result else ""

_VARIANT_INHERITANCES = [
    "maternally inherited",
    "parternally inherited",
    "de novo",
    "?de novo (if not confirmed)",
    "unknown",
    "maternally and paternally inherited homozygous"
]

def _variant_inheritance_validator(x: Any) -> Tuple[bool, str]:
    result = x in _VARIANT_INHERITANCES
    return result, f"Value is not one of {_VARIANT_INHERITANCES}" if not result else ""


_STUDY_TYPES = ["case report", "case series", "cohort analysis", None]


def _study_type_validator(x: Any) -> Tuple[bool, str]:
    result = x in _STUDY_TYPES
    return result, f"Value is not one of {_STUDY_TYPES}" if not result else ""


_FUNCTIONAL_STUDY_TYPES = [
    "functional studies using patient cells supports (or lacks) pathogenicity of the variant",
    "functional studies using animal cells or model supports (or lacks) pathogenicity of the variant",
    "functional studies using a cell line supports (or lacks) pathogenicity of the variant",
    "in silico analysis supports (or lacks) pathogenicity of the variant",
    "none",
]


def _functional_study_validator(x: Any) -> Tuple[bool, str]:
    result = x in _FUNCTIONAL_STUDY_TYPES
    return result, f"Value is not one of {_FUNCTIONAL_STUDY_TYPES}" if not result else ""


def _yes_no_validator(x: Any) -> Tuple[bool, str]:
    result = x in ["Y", "N"]
    return result, "Value is not one of 'Y' or 'N'" if not result else ""

# TODO: text description for phenotype can be empty, HPO list can be empty
# TODO: need variant_inheritance validator
# TODO: need individual id validator

VALIDATORS = {
    "paper_build": _paper_build_validator,
    "paper_variant": _not_none_validator,
    "individual_id": _not_none_validator,
    "pheno_text_description": _not_none_validator,
    "phenotype": _phenotype_validator,
    "transcript": _transcript_validator,
    "hgvsc": _hgvs_c_validator,
    "hgvsp": _hgvs_p_validator,
    "variant_type": _variant_type_validator,
    "zygosity": _zygosity_validator,
    "variant_inheritance": _variant_inheritance_validator,
    "pmid": _is_digit_validator,
    "author": _not_none_validator,
    "year": _is_digit_validator,
    "journal": _not_none_validator,
    "study_type": _study_type_validator,
    "functional_study": _functional_study_validator,
    "in_supplement": _yes_no_validator,
}

# %% Read in the truth spreadsheet.

wb = openpyxl.load_workbook(TRUTH_XLSX)

# %% Extract the gene names, evidence bases, and pmids from the PAPERS_SHEET_NAME sheet.

genes: List[Tuple[str, str, List[str]]] = []

papers_sheet = wb[PAPERS_SHEET_NAME]


def _convert_to_str(val: Any) -> str:
    return str(int(val)) if val.is_integer() else str(val)


for row in papers_sheet.iter_rows(min_row=2):
    gene = row[0].value
    if gene:
        pmids: List[str] = []
        pmid = row[5]
        while pmid.value:
            if str(pmid.value).replace(".", "").isdigit():
                pmids.append(_convert_to_str(pmid.value))
            else:
                print(f"WARNING: Skipping non-numeric PMID: {pmid.value} for gene {gene}")
            pmid = pmid.offset(row=0, column=1)

        genes.append((str(gene), str(row[1].value), pmids))

# Also make a gene dataframe with the gene and evidence base.
gene_df = pd.DataFrame([(gene, evidence_base) for gene, evidence_base, _ in genes], columns=["gene", "evidence_base"])

# Assign genes to train or test based on train_frac, drawn evenly from each evidence_base.
gene_df["is_train"] = (
    gene_df.groupby("evidence_base").cumcount() < gene_df.groupby("evidence_base").cumcount().max() * TRAIN_FRAC
)

# Convert the list of tuples to a dataframe where every row corresponds to a gene-paper pair and the columns are gene,
# evidence_base, and pmid.
gene_paper_df = pd.DataFrame(
    [(gene, pmid) for gene, _, pmids in genes for pmid in pmids],
    columns=["gene", "pmid"],
)


# %% For every gene, find the correct sheet and extract the evidence.

evidence_df = pd.DataFrame(columns=list(CONTENT_SHEET_COLUMN_HEADINGS.values()))

for gene_tuple in genes:
    gene_name = gene_tuple[0]

    ws = next((ws for ws in wb.worksheets if ws.title.startswith(gene_name)), None)

    if ws is None:
        print(f"WARNING: No sheet found for gene {gene_name}")
        continue

    # TODO: consider empty rows within a worksheet, not just as the last one.

    # Restrict the worksheet to only include rows and columns that contain data.
    # We're going to stop at the first empty row and delete the rest.
    row_found = False
    for _row_index, row in enumerate(ws.iter_rows()):
        if not any(cell.value for cell in row):
            row_found = True
            break

    # Delete this row and all that follow
    if row_found:
        ws.delete_rows(_row_index + 1, ws.max_row - _row_index)

    col_found = False
    for _col_index, col in enumerate(ws.iter_cols()):
        if not any(cell.value for cell in col):
            col_found = True
            break

    # Delete this column and all that follow
    if col_found:
        ws.delete_cols(_col_index + 1, ws.max_column - _col_index)

    # Make sure the set of column names in the worksheet matches the expected set.
    expected_headings = set(CONTENT_SHEET_COLUMN_HEADINGS.keys())

    if {str(cell.value) for cell in ws[1]} != expected_headings:
        print(f"WARNING: Column headings do not match for gene {gene_name} / sheet {ws.title}")
        ws_headings = {str(cell.value) for cell in ws[1]}
        print(f"  The following headings were not found in the worksheet: {expected_headings - ws_headings}")
        print(
            f"  The following headings were found in the worksheet but not expected: {ws_headings - expected_headings}"
        )
        continue

    # Extract the evidence from this sheet and add it to the evidence dataframe.

    # Determine the column index for each of the expected headings.
    col_indices = {heading.value: col_index - 1 for col_index, heading in enumerate(ws[1], start=1)}

    # Extract the evidence from the worksheet.
    gene_evidence_raw = []
    for row in ws.iter_rows(min_row=2):
        row_dict = {"gene": gene_name}
        for k, v in CONTENT_SHEET_COLUMN_HEADINGS.items():
            if str(row[col_indices[k]].value).replace(".", "").isdigit() and v in FORCED_STRING_FIELDS:
                row_dict[v] = str(int(row[col_indices[k]].value))  # type: ignore
            else:
                row_dict[v] = row[col_indices[k]].value  # type: ignore
        gene_evidence_raw.append(row_dict)

    gene_evidence_df = pd.DataFrame(gene_evidence_raw, columns=["gene"] + list(CONTENT_SHEET_COLUMN_HEADINGS.values()))

    # Do some more hygiene checks.
    if not any(gene_evidence_df["paper_variant"]):
        print(f"WARNING: No variants found in manual evidence for gene {gene_name} / sheet {ws.title}, skipping gene.")
        continue

    # Column-wise checks.
    gene_evidence_df["drop"] = False

    for index, row in enumerate(gene_evidence_df.itertuples()):
        if not row.paper_variant:
            print(
                f"WARNING: gene {gene_name} / sheet {ws.title} SKIPPING ROW {index+2} , "
                "due to None in paper_variant."
            )
            gene_evidence_df.at[index, "drop"] = True
            continue

        failed_cols = []

        for col in gene_evidence_df.columns:
            if col in VALIDATORS:
                validation_result, validation_info = VALIDATORS[col](getattr(row, col))
                if not validation_result:
                    failed_cols.append(col)

        if failed_cols:
            print(
                f"WARNING: gene {gene_name} / sheet {ws.title} SKIPPING ROW {index+2} , "
                f"due to failed validations: {failed_cols}"
            )
            gene_evidence_df.at[index, "drop"] = True

    gene_evidence_df = gene_evidence_df[~gene_evidence_df["drop"]]
    gene_evidence_df.drop(columns=["drop"], inplace=True)

    if gene_evidence_df.empty:
        print(f"WARNING: No valid rows found for gene {gene_name}, skipping gene.")
        continue

    print(f"## INFO: Adding {len(gene_evidence_df)} rows of evidence for gene {gene_name}")
    evidence_df = pd.concat([evidence_df, gene_evidence_df])


# %% Post-process evidence_df before writing out.

# Move the "gene" column to the front.
evidence_df = evidence_df[["gene"] + [col for col in evidence_df.columns if col != "gene"]]

# Add a paper_id column right after pmid that is formatted as "pmid:{pmid}".
evidence_df["paper_id"] = "pmid:" + evidence_df["pmid"].astype(str)

# Helper function for getting paper titles.
ncbi_client = NcbiLookupClient(
    CosmosCachingWebClient(get_dotenv_settings(filter_prefix="EVAGG_CONTENT_CACHE_")),
    get_dotenv_settings(filter_prefix="NCBI_EUTILS_"),
)


@cache
def get_paper(pmid: str) -> Any:
    return ncbi_client.fetch(pmid)


def get_paper_title(pmid: str) -> str:
    if paper := get_paper(pmid):
        return paper.props.get("title", "unknown")
    return "unknown"


def get_pmc_oa(pmid: str) -> bool:
    if paper := get_paper(pmid):
        return paper.props.get("is_pmc_oa", False)
    return False


def get_license(pmid: str) -> str | None:
    if paper := get_paper(pmid):
        return paper.props.get("license", None)
    return None


def get_pmcid(pmid: str) -> str | None:
    if paper := get_paper(pmid):
        return paper.props.get("pmcid", None)
    return None


# Now get the paper title.
evidence_df["paper_title"] = evidence_df["pmid"].apply(get_paper_title)

# Now add the pmc_oa status.
evidence_df["is_pmc_oa"] = evidence_df["pmid"].apply(get_pmc_oa)

# Now add the license.
evidence_df["license"] = evidence_df["pmid"].apply(get_license)

# And the pmcid.
evidence_df["pmcid"] = evidence_df["pmid"].apply(get_pmcid)

# And the link, which is just a link to the paper on pubmed.
evidence_df["link"] = "https://pubmed.ncbi.nlm.nih.gov/" + evidence_df["pmid"].astype(str) + "/"


# %% Write gene_paper_df and evidence_df to disk, splitting into two files each based on is_train from gene_df.

os.makedirs(OUTPUT_ROOT, exist_ok=True)

gene_paper_df_train = gene_paper_df.merge(gene_df, on="gene", how="left").query("is_train")
gene_paper_df_test = gene_paper_df.merge(gene_df, on="gene", how="left").query("~is_train")

gene_paper_df_train.drop(columns=["evidence_base", "is_train"], inplace=True)
gene_paper_df_test.drop(columns=["evidence_base", "is_train"], inplace=True)

gene_paper_df_train.to_csv(os.path.join(OUTPUT_ROOT, "papers_train_v1.tsv"), sep="\t", index=False)
gene_paper_df_test.to_csv(os.path.join(OUTPUT_ROOT, "papers_test_v1.tsv"), sep="\t", index=False)

evidence_df_train = evidence_df.merge(gene_df, on="gene", how="left").query("is_train")
evidence_df_test = evidence_df.merge(gene_df, on="gene", how="left").query("~is_train")

evidence_df_train.drop(columns=["evidence_base", "is_train"], inplace=True)
evidence_df_test.drop(columns=["evidence_base", "is_train"], inplace=True)

evidence_df_train.to_csv(os.path.join(OUTPUT_ROOT, "evidence_train_v1.tsv"), sep="\t", index=False)
evidence_df_test.to_csv(os.path.join(OUTPUT_ROOT, "evidence_test_v1.tsv"), sep="\t", index=False)

# %%
